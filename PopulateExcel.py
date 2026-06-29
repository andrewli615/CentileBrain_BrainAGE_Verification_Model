#!/usr/bin/env python3
"""
Populate brainAGE_template_*.xlsx files from the matching ethnoracial CSV files.

Typical use:
    python3 PopulateExcel.py --dataset self_chinese --write
    python3 PopulateExcel.py --dataset self_mexico --dry-run
    python3 PopulateExcel.py --all --write

To run by editing only a few keywords, change the DEFAULT_* values below and run:
    python3 PopulateExcel.py
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
ETHNORACIAL_DIR = ROOT / "ethnoracial_data"

# Edit these keywords when you want to run one dataset without command-line args.
DEFAULT_DATASET = "self_black"
DEFAULT_CSV_SUBFOLDER = ""
DEFAULT_WRITE_FILES = False
GENERATE_SUBJECT_ID_IF_MISSING = True
ALL_DATASET_PREFIX = "self_"
STRICT_ROW_KEY_MATCH = True
ALLOW_ROW_ORDER_MISMATCH_DATASETS = {"self_chinese"}

SEX_KEYWORDS = ("male", "female")
MODALITY_ORDER = ("thickness", "area", "volume")
DEMOGRAPHIC_COLUMNS = {
    "SITE",
    "SubjectID",
    "age",
    "sex",
    "ScannerType",
    "FreeSurfer_Version",
}

SUBCORTICAL_MAP = {
    "Lthal": "LeftThalamus",
    "Rthal": "RightThalamus",
    "Lcaud": "LeftCaudate",
    "Rcaud": "RightCaudate",
    "Lput": "LeftPutamen",
    "Rput": "RightPutamen",
    "Lpal": "LeftPallidum",
    "Rpal": "RightPallidum",
    "Lhippo": "LeftHippocampus",
    "Rhippo": "RightHippocampus",
    "Lamyg": "LeftAmygdala",
    "Ramyg": "RightAmygdala",
    "Laccumb": "LeftAccumbensarea",
    "Raccumb": "RightAccumbensarea",
}

CSV_REGION_ALIASES = {
    "entorhil": "entorhinal",
    "supramargil": "supramarginal",
}


@dataclass(frozen=True)
class TemplateTarget:
    path: Path
    sex: str
    age_group: str


def normalize_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def detect_sex(path: Path) -> str | None:
    stem = path.stem.casefold()
    if "female" in stem:
        return "female"
    if "male" in stem:
        return "male"
    return None


def detect_modality(path: Path, columns: list[str]) -> str | None:
    stem = path.stem.casefold()
    normalized_cols = {normalize_text(col) for col in columns}

    if "subcortical" in stem or "volume" in stem or {"leftthalamus", "lthal"} & normalized_cols:
        return "volume"
    if "meanthickness" in columns or "thickness" in stem or "cortical" in stem:
        return "thickness"
    if "meanarea" in columns or "area" in stem or "surface" in stem:
        return "area"
    return None


def detect_age_group(path: Path) -> str | None:
    name = path.name.casefold()
    if "<=40" in name or "≤40" in name or "le40" in name:
        return "<=40"
    if ">40" in name or "gt40" in name:
        return ">40"
    return None


def age_mask(ages: pd.Series, age_group: str) -> pd.Series:
    numeric_age = pd.to_numeric(ages, errors="coerce")
    if age_group == "<=40":
        return numeric_age <= 40
    if age_group == ">40":
        return numeric_age > 40
    raise ValueError(f"Unsupported age group: {age_group}")


def classify_csvs(csv_dir: Path) -> dict[str, dict[str, Path]]:
    classified: dict[str, dict[str, Path]] = {sex: {} for sex in SEX_KEYWORDS}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        columns = pd.read_csv(csv_path, nrows=0).columns.tolist()
        sex = detect_sex(csv_path)
        modality = detect_modality(csv_path, columns)
        if sex is None or modality is None:
            print(f"Skipping unclassified CSV: {csv_path.relative_to(ROOT)}")
            continue
        if modality in classified[sex]:
            previous = classified[sex][modality]
            raise ValueError(
                f"Found more than one {sex} {modality} CSV: "
                f"{previous.relative_to(ROOT)} and {csv_path.relative_to(ROOT)}"
            )
        classified[sex][modality] = csv_path
    return classified


def find_templates(template_dir: Path) -> list[TemplateTarget]:
    targets: list[TemplateTarget] = []
    for path in sorted(template_dir.glob("brainAGE_template_*.xlsx")):
        sex = detect_sex(path)
        age_group = detect_age_group(path)
        if sex is None or age_group is None:
            print(f"Skipping unclassified template: {path.relative_to(ROOT)}")
            continue
        targets.append(TemplateTarget(path=path, sex=sex, age_group=age_group))
    return targets


def make_row_key(df: pd.DataFrame) -> pd.Series:
    key_columns = [col for col in ("SubjectID", "SITE", "age") if col in df.columns]
    if not key_columns:
        return pd.Series(df.index.astype(str), index=df.index)
    return df[key_columns].astype(str).agg("|".join, axis=1)


def load_sex_data(csvs_for_sex: dict[str, Path], strict_row_key_match: bool) -> pd.DataFrame:
    missing = [modality for modality in MODALITY_ORDER if modality not in csvs_for_sex]
    if missing:
        raise ValueError(f"Missing required CSV modality/modalities: {', '.join(missing)}")

    frames: dict[str, pd.DataFrame] = {
        modality: pd.read_csv(csvs_for_sex[modality]) for modality in MODALITY_ORDER
    }
    base = pd.DataFrame(index=frames["thickness"].index)
    base_key = make_row_key(frames["thickness"])

    for modality, frame in frames.items():
        if len(frame) != len(base):
            raise ValueError(
                f"{csvs_for_sex[modality].name} has {len(frame)} rows, "
                f"but {csvs_for_sex['thickness'].name} has {len(base)} rows"
            )
        frame_key = make_row_key(frame)
        if not base_key.equals(frame_key):
            message = (
                f"{csvs_for_sex[modality].name} does not perfectly match "
                "the thickness row keys. Without a shared SubjectID, combining rows by "
                "row order may attach measurements from different participants."
            )
            if strict_row_key_match:
                raise ValueError(message)
            print(f"Warning: {message} Rows will be combined by row order.")

    output_columns: list[pd.Series] = []
    seen_direct_columns: set[str] = set()

    for modality, frame in frames.items():
        for col in frame.columns:
            if col in DEMOGRAPHIC_COLUMNS and col not in seen_direct_columns:
                output_columns.append(frame[col].rename(col))
                seen_direct_columns.add(col)
            elif col in DEMOGRAPHIC_COLUMNS:
                continue

            namespaced_col = f"{modality}::{col}"
            output_columns.append(frame[col].rename(namespaced_col))
            if col not in seen_direct_columns:
                output_columns.append(frame[col].rename(col))
                seen_direct_columns.add(col)

    return pd.concat(output_columns, axis=1).copy()


def direct_lookup(row: pd.Series, header: str) -> Any:
    if header in row.index:
        return row[header]
    normalized_header = normalize_text(header)
    for col in row.index:
        if normalize_text(col) == normalized_header:
            return row[col]
    return None


def template_header_to_csv_column(header: str) -> tuple[str, str] | None:
    subcortical = SUBCORTICAL_MAP.get(header)
    if subcortical:
        return "volume", subcortical

    match = re.match(r"^([LR])_(.+)_(thickavg|surfavg)$", header)
    if not match:
        return None

    side, region, measure = match.groups()
    source_region = CSV_REGION_ALIASES.get(region, region)
    source_side = "lh" if side == "L" else "rh"
    source = f"{source_side}_{source_region}"
    if measure == "surfavg":
        return "area", source
    return "thickness", source


def generated_subject_id(row: pd.Series, output_index: int) -> str:
    site = row.get("SITE")
    site_text = str(site) if pd.notna(site) and str(site).strip() else "participant"
    return f"{site_text}_{output_index:05d}"


def value_for_header(row: pd.Series, header: str, sex: str, output_index: int) -> Any:
    value = direct_lookup(row, header)
    if value is not None:
        return clean_value(value)

    mapped_column = template_header_to_csv_column(header)
    if mapped_column:
        modality, csv_column = mapped_column
        value = direct_lookup(row, f"{modality}::{csv_column}")
        if value is not None:
            return clean_value(value)
        value = direct_lookup(row, csv_column)
        if value is not None:
            return clean_value(value)

    if header == "sex":
        return 1 if sex == "male" else 2
    if header == "SubjectID" and GENERATE_SUBJECT_ID_IF_MISSING:
        return generated_subject_id(row, output_index)
    return None


def clean_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def clear_existing_data(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def populate_template(target: TemplateTarget, data: pd.DataFrame, write_files: bool) -> int:
    filtered = data.loc[age_mask(data["age"], target.age_group)].reset_index(drop=True)
    wb = load_workbook(target.path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    if write_files:
        clear_existing_data(ws)
        for row_number, (_, row) in enumerate(filtered.iterrows(), start=2):
            output_index = row_number - 1
            for col_number, header in enumerate(headers, start=1):
                if header is None:
                    continue
                ws.cell(row=row_number, column=col_number, value=value_for_header(row, str(header), target.sex, output_index))
        wb.save(target.path)

    return len(filtered)


def run_dataset(dataset: str, csv_subfolder: str, write_files: bool, strict_row_key_match: bool) -> None:
    dataset_dir = ETHNORACIAL_DIR / dataset
    if not dataset_dir.exists():
        raise FileNotFoundError(f"Dataset folder not found: {dataset_dir}")

    csv_dir = dataset_dir / csv_subfolder if csv_subfolder else dataset_dir
    template_dir = dataset_dir
    classified = classify_csvs(csv_dir)
    targets = find_templates(template_dir)

    if not targets:
        print(f"{dataset}: no brainAGE_template_*.xlsx files found; nothing to populate.")
        return

    print(f"\nDataset: {dataset}")
    print(f"CSV folder: {csv_dir.relative_to(ROOT)}")
    print("Mode:", "WRITE" if write_files else "DRY RUN")

    dataset_strict_row_key_match = (
        strict_row_key_match and dataset not in ALLOW_ROW_ORDER_MISMATCH_DATASETS
    )

    data_by_sex: dict[str, pd.DataFrame] = {}
    for target in targets:
        if target.sex not in data_by_sex:
            data_by_sex[target.sex] = load_sex_data(
                classified[target.sex], dataset_strict_row_key_match
            )
        count = populate_template(target, data_by_sex[target.sex], write_files)
        print(f"  {target.path.name}: {count} rows")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset", default=DEFAULT_DATASET, help="Folder under ethnoracial_data to process.")
    parser.add_argument("--csv-subfolder", default=DEFAULT_CSV_SUBFOLDER, help="Optional subfolder containing the six CSV inputs.")
    parser.add_argument(
        "--all",
        action="store_true",
        help=f"Process every {ALL_DATASET_PREFIX} ethnoracial_data subfolder that has templates.",
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--write", action="store_true", help="Write populated rows into the template workbooks.")
    mode.add_argument("--dry-run", action="store_true", help="Show row counts without editing workbooks.")
    parser.add_argument(
        "--allow-row-order-mismatch",
        action="store_true",
        help="Allow CSVs with mismatched row keys to be combined by row order.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    write_files = args.write or (DEFAULT_WRITE_FILES and not args.dry_run)
    strict_row_key_match = STRICT_ROW_KEY_MATCH and not args.allow_row_order_mismatch

    if args.all:
        for dataset_dir in sorted(path for path in ETHNORACIAL_DIR.iterdir() if path.is_dir()):
            if dataset_dir.name.startswith(ALL_DATASET_PREFIX) and list(dataset_dir.glob("brainAGE_template_*.xlsx")):
                run_dataset(dataset_dir.name, "", write_files, strict_row_key_match)
    else:
        run_dataset(args.dataset, args.csv_subfolder, write_files, strict_row_key_match)

    if not write_files:
        print("\nDry run only. Add --write or set DEFAULT_WRITE_FILES = True to populate files.")


if __name__ == "__main__":
    main()
